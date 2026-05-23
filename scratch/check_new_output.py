import openpyxl

fn = "test_mapped_out.xlsx"
try:
    wb = openpyxl.load_workbook(fn, data_only=True)
    print(f"\n--- Checking Mapped Test Output: {fn} (eval values) ---")
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        div_cells = []
        for r in range(1, min(sheet.max_row + 1, 300)):
            for c in range(1, min(sheet.max_column + 1, 150)):
                val = sheet.cell(row=r, column=c).value
                if val == "#DIV/0!":
                    div_cells.append((r, c))
        if div_cells:
            print(f"Sheet '{sheetname}' has {len(div_cells)} #DIV/0! cells.")
            wb_form = openpyxl.load_workbook(fn, data_only=False)
            sheet_form = wb_form[sheetname]
            for r, c in div_cells[:15]:
                cell_letter = openpyxl.utils.get_column_letter(c)
                print(f"  {cell_letter}{r}: evaluated to #DIV/0!, Formula: {sheet_form.cell(row=r, column=c).value}")
        else:
            print(f"Sheet '{sheetname}' has no #DIV/0! cells.")
except Exception as e:
    print(f"Error checking mapped test output: {e}")
