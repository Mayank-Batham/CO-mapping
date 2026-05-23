import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"Sheet: {name}, dimensions: {sheet.dimensions}")
    # Let's search for formulas that might contain division or references
    # Particularly in summary rows (e.g. 86-100 in Theory)
    div_formulas = []
    for r in range(1, min(sheet.max_row + 1, 150)):
        for c in range(1, min(sheet.max_column + 1, 130)):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                # Print formula if it has division or looks like average/percentage
                if "/" in val or "AVERAGE" in val or "COUNT" in val:
                    div_formulas.append((r, c, val))
    print(f"Found {len(div_formulas)} potential formulas in {name}. Showing first 15:")
    for r, c, val in div_formulas[:15]:
        print(f"  Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
