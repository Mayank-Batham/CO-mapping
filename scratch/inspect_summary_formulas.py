import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["Theory"]
print("--- Theory summary formulas at rows 85-95 ---")
for r in range(85, 96):
    row_formulas = {}
    for c in range(1, 40): # columns A to AM
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str) and val.startswith("="):
            row_formulas[openpyxl.utils.get_column_letter(c)] = val
    if row_formulas:
        print(f"Row {r} has formulas in cols: {list(row_formulas.items())[:5]}...")
