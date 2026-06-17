import openpyxl

def extract_lab_marks_from_excel(file_path, num_marks_cols):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Find active sheet or sheet containing 'lab'
    sheet = None
    for name in wb.sheetnames:
        if 'lab' in name.lower():
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb.active
        
    print(f"Reading sheet: {sheet.title} from {file_path}")
    
    # Scan for header row containing "USN"
    header_row_idx = None
    usn_col_idx = None
    name_col_idx = None
    
    for r in range(1, 30):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
        row_vals_str = [str(v).strip().upper() if v is not None else "" for v in row_vals]
        if "USN" in row_vals_str:
            header_row_idx = r
            usn_col_idx = row_vals_str.index("USN") + 1
            # Find Name column
            for idx, val in enumerate(row_vals_str):
                if "NAME" in val or "STUDENT" in val:
                    name_col_idx = idx + 1
                    break
            if not name_col_idx:
                name_col_idx = usn_col_idx + 1
            break
            
    if not header_row_idx:
        header_row_idx = 8
        usn_col_idx = 2
        name_col_idx = 3
        
    print(f"Found header at row {header_row_idx}: USN col = {usn_col_idx}, Name col = {name_col_idx}")
    
    students_data = {}
    
    max_row = sheet.max_row
    for r in range(header_row_idx + 1, max_row + 1):
        usn = sheet.cell(row=r, column=usn_col_idx).value
        name = sheet.cell(row=r, column=name_col_idx).value
        
        if usn is not None:
            usn_str = str(usn).strip()
            if not usn_str or usn_str.upper() in ("USN", "NONE", "NAN") or len(usn_str) < 5:
                continue
            
            # Read marks (columns 4 up to 4 + num_marks_cols)
            marks = []
            for m_col in range(4, 4 + num_marks_cols):
                val = sheet.cell(row=r, column=m_col).value
                if val is not None and str(val).strip() != "":
                    try:
                        marks.append(float(val))
                    except ValueError:
                        marks.append(None)
                else:
                    marks.append(None)
                    
            see_100 = None
            if num_marks_cols == 13: # Stand Alone Lab has Column 19 (S)
                val_see = sheet.cell(row=r, column=19).value
                if val_see is not None and str(val_see).strip() != "":
                    try:
                        see_100 = float(val_see)
                    except ValueError:
                        pass
                        
            name_str = str(name).strip() if name is not None else ""
            students_data[usn_str.upper()] = {
                'name': name_str,
                'marks': marks,
                'see_100': see_100
            }
            
    return students_data

# Test Standalone Lab file
print("\n--- STANDALONE LAB ---")
standalone_data = extract_lab_marks_from_excel("STAND ALONE LAB MARKS.xlsx", 13)
print(f"Extracted {len(standalone_data)} students.")
# Print first 3 students
for k, v in list(standalone_data.items())[:3]:
    print(f"{k}: Name='{v['name']}', Marks={v['marks']}, SEE_100={v['see_100']}")

# Test IPCC Lab file
print("\n--- IPCC LAB ---")
ipcc_data = extract_lab_marks_from_excel("IPCC Lab component.xlsx", 6)
print(f"Extracted {len(ipcc_data)} students.")
# Print first 3 students
for k, v in list(ipcc_data.items())[:3]:
    print(f"{k}: Name='{v['name']}', Marks={v['marks']}, SEE_100={v['see_100']}")
