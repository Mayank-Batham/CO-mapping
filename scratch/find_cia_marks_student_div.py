import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["CIA MARKS"]
print("--- Division formulas in CIA MARKS student rows (14-83) ---")
for r in range(14, 84):
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str) and val.startswith("="):
            if "/" in val:
                print(f"Row {r} Col {openpyxl.utils.get_column_letter(c)}: {val}")
