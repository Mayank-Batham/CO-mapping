import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["Theory"]
empty_max_cols = []
for c in range(4, 102): # Col D to CW
    val = sheet.cell(row=14, column=c).value
    if val is None or val == 0 or val == "":
        empty_max_cols.append(openpyxl.utils.get_column_letter(c))
print(f"Columns in Theory sheet with empty or zero max marks in row 14 ({len(empty_max_cols)} columns):")
print(empty_max_cols)
