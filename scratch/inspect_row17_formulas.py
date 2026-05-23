import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["Theory"]
for c in range(1, sheet.max_column + 1):
    val = sheet.cell(row=17, column=c).value
    if isinstance(val, str) and val.startswith("="):
        col_letter = openpyxl.utils.get_column_letter(c)
        print(f"Col {c} ({col_letter}): {val}")
