import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["Theory"]
print("--- Formulas in Theory sheet rows 17-85 ---")
for r in range(17, 86):
    row_formulas = {}
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str) and val.startswith("="):
            row_formulas[openpyxl.utils.get_column_letter(c)] = val
    if row_formulas:
        print(f"Row {r} has formulas: {list(row_formulas.items())[:5]}... (total {len(row_formulas)} formulas)")
        break # They are likely identical for all rows, so just show the first one
else:
    print("No formulas found in rows 17-85 of Theory sheet.")
