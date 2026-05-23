import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["Theory"]
for col in [100, 101]: # CV and CW
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"\n--- COLUMN {col_letter} ({col}) ---")
    print(f"Row 14 (Max Marks):", sheet.cell(row=14, column=col).value)
    print(f"Row 15:", sheet.cell(row=15, column=col).value)
    print(f"Row 86 (COUNTIFS):", sheet.cell(row=86, column=col).value)
    print(f"Row 87 (SUMIF):", sheet.cell(row=87, column=col).value)
    print(f"Row 88 (Formula):", sheet.cell(row=88, column=col).value)
