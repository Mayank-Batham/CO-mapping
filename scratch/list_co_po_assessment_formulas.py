import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["CO_PO_ASSESSMENT"]
print("--- CO_PO_ASSESSMENT formulas ---")
count = 0
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str) and val.startswith("="):
            count += 1
            if count <= 25:
                print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
print(f"Total formulas: {count}")
