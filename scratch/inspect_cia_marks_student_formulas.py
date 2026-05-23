import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["CIA MARKS"]
print("--- Formulas in CIA MARKS student rows (14-83) ---")
for r in range(14, 84):
    row_formulas = {}
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str) and val.startswith("="):
            row_formulas[openpyxl.utils.get_column_letter(c)] = val
    if row_formulas:
        print(f"Row {r} has formulas in cols: {list(row_formulas.items())[:8]}...")
        break
