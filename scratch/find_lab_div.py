import openpyxl

wb = openpyxl.load_workbook("STAND ALONE THEORY EMPTY TEMPLATE.xlsx", data_only=False)
sheet = wb["Lab"]
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str) and val.startswith("="):
            if "/" in val:
                print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
