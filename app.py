import streamlit as st
import pandas as pd
import docx
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import io
import re
import os

# Load environment variables manually from .env file
script_dir = os.path.dirname(os.path.abspath(__file__))
script_env = os.path.join(script_dir, '.env')
cwd_env = '.env'

def load_env_file(path):
    if os.path.exists(path):
        with open(path, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    if '=' in line:
                        key, value = line.split('=', 1)
                        os.environ[key.strip()] = value.strip()

# Load project script directory .env first, then let CWD override if it differs
load_env_file(script_env)
if os.path.abspath(cwd_env) != os.path.abspath(script_env):
    load_env_file(cwd_env)


st.set_page_config(
    page_title="CIA Marks & CO Mapper",
    page_icon="📊",
    layout="wide"
)

# Custom Premium Styling
st.markdown("""
<style>
    /* Premium font styling and background enhancements */
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif;
    }
    
    .main-title {
        background: linear-gradient(135deg, #6C63FF 0%, #3F3D56 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 3rem;
        font-weight: 800;
        margin-bottom: 0.5rem;
    }
    
    .subtitle {
        color: #707070;
        font-size: 1.2rem;
        margin-bottom: 2rem;
    }
    
    .card {
        background-color: #f8fafc;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.05);
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        border: 1px solid #e2e8f0;
    }
    
    .card h3 {
        color: #0f172a !important;
        margin-top: 0px !important;
        margin-bottom: 8px !important;
        font-weight: 700 !important;
        font-size: 1.25rem !important;
    }
    
    .card p {
        color: #475569 !important;
        font-size: 0.95rem !important;
        line-height: 1.5 !important;
        margin-bottom: 0px !important;
    }
    
    .tag-coming-soon {
        background: linear-gradient(135deg, #FF6B6B 0%, #FF8E53 100%);
        color: white;
        padding: 3px 8px;
        border-radius: 6px;
        font-size: 0.75rem;
        font-weight: bold;
        display: inline-block;
        margin-left: 10px;
    }
    
    .tag-active {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%);
        color: white;
        padding: 3px 8px;
        border-radius: 6px;
        font-size: 0.75rem;
        font-weight: bold;
        display: inline-block;
        margin-left: 10px;
    }
    
    /* Sleek buttons */
    .stButton>button {
        background: linear-gradient(135deg, #6C63FF 0%, #514DFD 100%);
        color: white;
        border-radius: 8px;
        padding: 0.6rem 2rem;
        font-weight: 600;
        border: none;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(108, 99, 255, 0.4);
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(108, 99, 255, 0.6);
        background: linear-gradient(135deg, #514DFD 0%, #6C63FF 100%);
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# ----------------- PARSING HELPER FUNCTIONS -----------------

def parse_xls_header(header_str):
    """Parses header strings like '7 (A)', '7 (B)', '7a', '1.0' into (q_num, part)."""
    header_str = str(header_str).strip()
    # Check for formats like "7 (A)", "7 (a)", "7A", "7a", "7 ( A )"
    m = re.match(r'^(\d+)\s*\(?\s*([a-zA-Z])\s*\)?', header_str)
    if m:
        q_num = int(m.group(1))
        part = m.group(2).lower()
        return q_num, part
    # Check for plain digit
    m = re.match(r'^(\d+)(?:\.0)?$', header_str)
    if m:
        q_num = int(m.group(1))
        return q_num, 'a'
    return None, None

def extract_cos_from_docx(file_stream):
    """Parses a DOCX file and extracts Course Outcome mappings."""
    document = docx.Document(file_stream)
    co_mapping = {} # (q_num, part) -> CO
    for table in document.tables:
        q_row_idx = None
        part_row_idx = None
        co_row_idx = None
        for i, row in enumerate(table.rows):
            cells = [cell.text.strip() for cell in row.cells]
            if not cells: continue
            if 'Question No.' in cells[0]:
                if q_row_idx is None:
                    q_row_idx = i
                elif part_row_idx is None:
                    part_row_idx = i
            if 'Course Outcome' in cells[0] or 'Course outcome' in cells[0]:
                co_row_idx = i
        if q_row_idx is not None and co_row_idx is not None:
            q_cells = [c.text.strip() for c in table.rows[q_row_idx].cells]
            if part_row_idx is not None:
                part_cells = [c.text.strip() for c in table.rows[part_row_idx].cells]
            else:
                part_cells = [""] * len(q_cells)
            co_cells = [c.text.strip() for c in table.rows[co_row_idx].cells]
            for idx in range(1, len(q_cells)):
                q = q_cells[idx]
                part_str = part_cells[idx].lower().strip() if idx < len(part_cells) else ""
                co = co_cells[idx].strip() if idx < len(co_cells) else ""
                if part_str not in ('a', 'b', 'c', 'd'):
                    part = 'a'
                else:
                    part = part_str
                if q and q.isdigit():
                    q_num = int(q)
                    co_val = co.replace("CO", "").strip()
                    if co_val: # Only map if the CO is not empty
                        co_mapping[(q_num, part)] = co_val
    return co_mapping

def extract_marks_from_xls(file_stream):
    """Parses IA Marks XLS generated and extracts marks per USN."""
    # Reading via xlrd for .xls support
    file_bytes = file_stream.read()
    file_stream.seek(0)
    df = pd.read_excel(io.BytesIO(file_bytes), engine='xlrd', header=None)
    header_row_idx = None
    usn_col = None
    for idx, row in df.iterrows():
        row_vals = [str(x).strip() for x in row.values]
        if 'USN' in row_vals:
            header_row_idx = idx
            usn_col = row_vals.index('USN')
            break
            
    if header_row_idx is None: 
        return {}
    
    headers = [str(x).strip() for x in df.iloc[header_row_idx].values]
    
    q_cols = {} # mapped (q_num, part) -> col id
    for col_idx, h in enumerate(headers):
        q_num, part = parse_xls_header(h)
        if q_num is not None:
            q_cols[(q_num, part)] = col_idx
            
    marks_data = {} # USN -> {(q_num, part): mark}
    for idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[idx]
        usn = str(row.iloc[usn_col]).strip()
        if not usn or usn.lower() in ('nan', 'none', '') or not usn.isalnum():
            continue
        
        st_marks = {}
        for (q_num, part), col_idx in q_cols.items():
            mark = row.iloc[col_idx]
            if pd.isna(mark) or str(mark).strip() in ['-', '', 'nan', 'A', 'None']:
                pass
            else:
                try:
                    st_marks[(q_num, part)] = float(mark)
                except ValueError:
                    pass
        marks_data[usn] = st_marks
        
    return marks_data

def extract_single_column_marks(file_stream):
    """Parses Quiz/AAT XLS and extracts single column marks (Theory Mark) per USN."""
    file_bytes = file_stream.read()
    file_stream.seek(0)
    df = pd.read_excel(io.BytesIO(file_bytes), engine='xlrd', header=None)
    header_row_idx = None
    usn_col = None
    marks_col = None
    
    for idx, row in df.iterrows():
        row_vals = [str(x).strip() for x in row.values]
        if 'USN' in row_vals:
            header_row_idx = idx
            usn_col = row_vals.index('USN')
            # Look for 'Theory Mark' or similar
            for col_idx, val in enumerate(row_vals):
                if 'theory mark' in val.lower() or 'marks' in val.lower() or 'mark' in val.lower():
                    if 'practical' not in val.lower():
                        marks_col = col_idx
                        break
            break
            
    if header_row_idx is None or usn_col is None:
        return {}
    if marks_col is None:
        marks_col = usn_col + 2 # fallback if not found directly
        
    marks_data = {}
    for idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[idx]
        usn = str(row.iloc[usn_col]).strip()
        if not usn or usn.lower() in ('nan', 'none', '') or not usn.isalnum():
            continue
        
        mark = row.iloc[marks_col]
        if pd.isna(mark) or str(mark).strip() in ['-', '', 'nan', 'A', 'None']:
            pass
        else:
            try:
                marks_data[usn] = float(mark)
            except ValueError:
                pass
                
    return marks_data

def extract_metadata(file_stream):
    """Extracts course metadata (Course Code, Faculty, etc.) from IA Marks file."""
    file_bytes = file_stream.read()
    file_stream.seek(0)
    df = pd.read_excel(io.BytesIO(file_bytes), engine='xlrd', header=None)
    metadata = {
        'faculty': '',
        'course_code': '',
        'course_title': '',
        'semester': '',
        'academic_year': '',
        'batch': ''
    }
    
    for idx, row in df.iterrows():
        row_str = [str(x) for x in row.values]
        for val in row_str:
            val_strip = val.strip()
            if ' - ' in val_strip and ('CIA' in val_strip or 'AIML' in val_strip):
                parts = val_strip.split(' - ')
                if len(parts) >= 3:
                    metadata['course_code'] = parts[1].strip()
                    metadata['course_title'] = parts[2].strip()
                    p0_parts = parts[0].split('-')
                    if len(p0_parts) >= 4:
                        metadata['semester'] = p0_parts[1].strip()
                        metadata['academic_year'] = f"{p0_parts[3].strip()}-{p0_parts[4].strip()}" if len(p0_parts) > 4 else p0_parts[3].strip()
            
            if 'staff name' in val_strip.lower() or 'faculty' in val_strip.lower():
                m = re.search(r'(?:staff name|faculty)\s*[:|-]+\s*(.*)', val_strip, re.IGNORECASE)
                if m:
                    metadata['faculty'] = m.group(1).strip()
                    
    if metadata['semester'] and metadata['academic_year']:
        try:
            sem = int(metadata['semester'])
            ay_start = int(metadata['academic_year'].split('-')[0])
            start_year = ay_start - ((sem + 1) // 2 - 1)
            end_year = start_year + 4
            metadata['batch'] = f"{start_year}-{str(end_year)[2:]}"
        except Exception as e:
            pass
            
    return metadata

# ----------------- CORE PROCESSING PIPELINE -----------------

def process_stand_alone_theory(template_file, qp_files, marks_files, quiz_file, aat_file):
    """Fills the Stand Alone Theory sheet cleanly and handles dynamic rosters."""
    wb = load_workbook(io.BytesIO(template_file.read()))
    sheet = wb['Theory'] if 'Theory' in wb.sheetnames else wb.active
    
    # 1. Configure Course Type Cells
    sheet.cell(row=7, column=4).value = "YES" # STAND ALONE THEORY
    sheet.cell(row=8, column=4).value = "NO"  # STAND ALONE LAB
    sheet.cell(row=9, column=4).value = "NO"  # IPCC
    
    # 2. Build Master Student Roster
    master_students = {}
    all_marks_files = []
    
    # Extract list of files
    if marks_files:
        all_marks_files.extend([f for f in marks_files.values() if f])
    if quiz_file:
        all_marks_files.append(quiz_file)
    if aat_file:
        all_marks_files.append(aat_file)
        
    for mf in all_marks_files:
        mf.seek(0)
        df = pd.read_excel(io.BytesIO(mf.read()), engine='xlrd', header=None)
        mf.seek(0)
        
        header_row_idx = None
        usn_col = None
        for idx, row in df.iterrows():
            row_vals = [str(x).strip() for x in row.values]
            if 'USN' in row_vals:
                header_row_idx = idx
                usn_col = row_vals.index('USN')
                break
                
        if header_row_idx is not None and usn_col is not None:
            row_vals = [str(x).strip() for x in df.iloc[header_row_idx].values]
            name_col = None
            for c_idx, val in enumerate(row_vals):
                if 'student' in val.lower() or 'name' in val.lower():
                    if 'staff' not in val.lower() and 'division' not in val.lower():
                        name_col = c_idx
                        break
            if name_col is None:
                name_col = usn_col + 1
                
            for idx in range(header_row_idx + 1, len(df)):
                row = df.iloc[idx]
                u = str(row.iloc[usn_col]).strip()
                n = str(row.iloc[name_col]).strip() if name_col < len(row) else ""
                if u and u.lower() not in ('nan', 'none', '') and u.isalnum():
                    master_students[u] = n
                    
    sorted_usns = sorted(master_students.keys())
    num_students = len(sorted_usns)
    
    if num_students == 0:
        raise ValueError("No students found in the uploaded marks files. Ensure Excel files are correct.")
        
    # 3. Populate Roster in Theory sheet (Rows 17 to 85)
    for i in range(85 - 17 + 1):
        r = 17 + i
        if i < num_students:
            usn = sorted_usns[i]
            sheet.cell(row=r, column=1).value = float(i + 1)
            sheet.cell(row=r, column=2).value = usn
            sheet.cell(row=r, column=3).value = master_students[usn]
        else:
            # Clear unused row columns A to DH (1 to 122)
            for c in range(1, 123):
                sheet.cell(row=r, column=c).value = None
                
    # 4. Clear all marks cells in student rows (Theory Cols D to CW, Rows 17 to 17 + N - 1)
    for r in range(17, 17 + num_students):
        for c in range(4, 102): # Col 4 (D) to 101 (CW)
            sheet.cell(row=r, column=c).value = None
            
    # 5. Clear unused student rows in other sheets
    if 'Lab' in wb.sheetnames:
        sheet_lab = wb['Lab']
        for r in range(9 + num_students, 78):
            for c in range(1, sheet_lab.max_column + 1):
                sheet_lab.cell(row=r, column=c).value = None
                
    if 'CIA MARKS' in wb.sheetnames:
        sheet_cia = wb['CIA MARKS']
        for r in range(14 + num_students, 84):
            for c in range(1, sheet_cia.max_column + 1):
                sheet_cia.cell(row=r, column=c).value = None
                
    # 6. Extract and write metadata from IA1 Marks if available
    metadata = {}
    for test_idx in [1, 2, 3]:
        if marks_files.get(test_idx):
            marks_files[test_idx].seek(0)
            metadata = extract_metadata(marks_files[test_idx])
            marks_files[test_idx].seek(0)
            if metadata['course_code']:
                break
                
    if metadata:
        if metadata['academic_year']:
            sheet.cell(row=4, column=1).value = f"Academic Year : {metadata['academic_year']}"
        if metadata['batch']:
            sheet.cell(row=4, column=4).value = f"Batch :{metadata['batch']}"
        if metadata['semester']:
            sheet.cell(row=4, column=17).value = f"Semester  : {metadata['semester']}"
        if metadata['course_code']:
            sheet.cell(row=5, column=1).value = f"COURSE  CODE : {metadata['course_code']}"
        if metadata['course_title']:
            sheet.cell(row=6, column=1).value = f"COURSE TITLE : {metadata['course_title']}"
        if metadata['faculty']:
            sheet.cell(row=10, column=1).value = f"Faculty : {metadata['faculty']}"
            
    # 7. Map Test Columns
    tests_col_map = {1: {}, 2: {}, 3: {}} 
    first_q1a_col = -1
    for col in range(1, 200):
        val = sheet.cell(row=16, column=col).value
        val_str = str(val).strip() if val else ""
        if val_str == 'Q1A':
            first_q1a_col = col
            break

    if first_q1a_col != -1:
        for test_idx in [1, 2, 3]:
            start_col = first_q1a_col + (test_idx - 1) * 32
            for q_num in range(1, 9):
                tests_col_map[test_idx][q_num] = start_col + (q_num - 1) * 4
                
    # 8. Write Test COs and Marks
    for test_idx in [1, 2, 3]:
        qp = qp_files.get(test_idx)
        marks = marks_files.get(test_idx)
        col_mappings = tests_col_map[test_idx]
        
        # Clear Row 12, 14, and 15 for CIA columns first to prevent old values
        for col_idx in col_mappings.values():
            for offset in range(4):
                sheet.cell(row=12, column=col_idx + offset).value = None
                sheet.cell(row=14, column=col_idx + offset).value = None
                sheet.cell(row=15, column=col_idx + offset).value = None
        
        # CO Mappings
        if qp:
            qp.seek(0)
            co_map = extract_cos_from_docx(qp)
            qp.seek(0)
            
            part_offsets = {'a': 0, 'b': 1, 'c': 2, 'd': 3}
            for (q_num, part), co_val in co_map.items():
                if q_num in col_mappings:
                    col_start = col_mappings[q_num]
                    target_col = col_start + part_offsets.get(part, 0)
                    sheet.cell(row=12, column=target_col).value = f"0,{co_val}"
                    
        # Marks Mappings
        if marks:
            marks.seek(0)
            st_marks = extract_marks_from_xls(marks)
            marks.seek(0)
            
            part_offsets = {'a': 0, 'b': 1, 'c': 2, 'd': 3}
            for i in range(num_students):
                r = 17 + i
                usn_val = sheet.cell(row=r, column=2).value
                if usn_val in st_marks:
                    for (q_num, part), mark in st_marks[usn_val].items():
                        if q_num in col_mappings:
                            col_start = col_mappings[q_num]
                            target_col = col_start + part_offsets.get(part, 0)
                            sheet.cell(row=r, column=target_col).value = mark

        # Set maximum marks and thresholds automatically based on part B activity
        # to eliminate division by zero errors in summaries!
        part_offsets = {'a': 0, 'b': 1, 'c': 2, 'd': 3}
        for q_num in range(1, 9):
            col_a = col_mappings[q_num]
            col_b = col_mappings[q_num] + 1
            
            # Check if part B is active (has marks mapped or CO mapped)
            part_b_active = False
            for i in range(num_students):
                r = 17 + i
                if sheet.cell(row=r, column=col_b).value is not None:
                    part_b_active = True
                    break
            if sheet.cell(row=12, column=col_b).value is not None:
                part_b_active = True
                
            if part_b_active:
                sheet.cell(row=14, column=col_a).value = 5.0
                sheet.cell(row=14, column=col_b).value = 5.0
                col_a_let = get_column_letter(col_a)
                col_b_let = get_column_letter(col_b)
                sheet.cell(row=15, column=col_a).value = f"=PRODUCT({col_a_let}14,0.65)"
                sheet.cell(row=15, column=col_b).value = f"=PRODUCT({col_b_let}14,0.65)"
            else:
                # Default to 10 marks for standard non-split questions
                sheet.cell(row=14, column=col_a).value = 10.0
                col_a_let = get_column_letter(col_a)
                sheet.cell(row=15, column=col_a).value = f"=PRODUCT({col_a_let}14,0.65)"
                            
    # 9. Map Quiz Marks (Column 101)
    if quiz_file:
        quiz_file.seek(0)
        quiz_marks = extract_single_column_marks(quiz_file)
        quiz_file.seek(0)
        for i in range(num_students):
            r = 17 + i
            usn_val = sheet.cell(row=r, column=2).value
            if usn_val in quiz_marks:
                sheet.cell(row=r, column=101).value = quiz_marks[usn_val]
                
    # 10. Map AAT Marks (Column 100)
    if aat_file:
        aat_file.seek(0)
        aat_marks = extract_single_column_marks(aat_file)
        aat_file.seek(0)
        for i in range(num_students):
            r = 17 + i
            usn_val = sheet.cell(row=r, column=2).value
            if usn_val in aat_marks:
                sheet.cell(row=r, column=100).value = aat_marks[usn_val]

    # 11. Prevent division by zero in Final SEE Grade (Col 108/DD) by filling active student rows with "NA"
    for i in range(num_students):
        r = 17 + i
        sheet.cell(row=r, column=108).value = "NA"

    # 12. Prevent division by zero in CES survey sheet by initializing row 11 with 0
    if 'CES' in wb.sheetnames:
        sheet_ces = wb['CES']
        for col in range(4, 14): # D to M
            if sheet_ces.cell(row=11, column=col).value is None:
                sheet_ces.cell(row=11, column=col).value = 0

    # 13. Prevent division by zero and fill CO-PO-PSO mapping matrix based on Course Code
    if 'CO_PO_PSO_MAPPING' in wb.sheetnames:
        sheet_cppm = wb['CO_PO_PSO_MAPPING']
        
        # Default initialization: fill all cells in CO1-CO10 (D4:R13) with 0 to prevent division by zero
        for r in range(4, 14):
            for c in range(4, 19): # Columns D to R
                sheet_cppm.cell(row=r, column=c).value = 0

        # Attempt to automatically load mapping based on course code
        if metadata and metadata.get('course_code'):
            target_code = re.sub(r'[^a-zA-Z0-9]', '', str(metadata['course_code'])).upper()
            
            # Extract core code by removing leading scheme indicators (e.g., "22", "21")
            m_target = re.match(r'^\d+(.*)', target_code)
            target_core = m_target.group(1) if m_target else target_code
            
            import os
            script_dir = os.path.dirname(os.path.abspath(__file__))
            src_mapping_file = None
            
            # Check current working directory first, then script directory
            for search_dir in ['.', script_dir]:
                try:
                    for f_name in os.listdir(search_dir):
                        f_upper = f_name.upper()
                        if 'CO' in f_upper and 'PO' in f_upper and 'MAPPING' in f_upper:
                            if f_upper.endswith('.XLSX') or f_upper.endswith('.XLS'):
                                src_mapping_file = os.path.join(search_dir, f_name)
                                break
                except:
                    pass
                if src_mapping_file:
                    break
                    
            if not src_mapping_file:
                for search_dir in ['.', script_dir]:
                    for fallback in ['2021-CO-PO -PSO-MAPPING.xlsx', '2021 CO PO PSO MAPPING.xls', '2021-CO-PO-PSO-MAPPING.xlsx', '2021-CO-PO-PSO-MAPPING.xls']:
                        fallback_path = os.path.join(search_dir, fallback)
                        if os.path.exists(fallback_path):
                            src_mapping_file = fallback_path
                            break
                    if src_mapping_file:
                        break
                
            if src_mapping_file:
                try:
                    engine = 'openpyxl' if src_mapping_file.lower().endswith('.xlsx') else 'xlrd'
                    raw_df = pd.read_excel(src_mapping_file, engine=engine, header=None)
                    
                    # Scan first 15 rows to find the header row containing 'Course Code'
                    header_row_idx = 0
                    course_code_col_idx = 0
                    found_header = False
                    
                    for r_idx in range(min(15, len(raw_df))):
                        row_vals = [str(x).strip().upper() for x in raw_df.iloc[r_idx].values]
                        for c_idx, val in enumerate(row_vals):
                            if ('COURSE' in val and 'CODE' in val) or ('SUBJ' in val and 'CODE' in val):
                                header_row_idx = r_idx
                                course_code_col_idx = c_idx
                                found_header = True
                                break
                        if found_header:
                            break
                            
                    if found_header:
                        cols = [str(x).strip() for x in raw_df.iloc[header_row_idx].values]
                        seen = {}
                        new_cols = []
                        for col in cols:
                            if col in seen:
                                seen[col] += 1
                                new_cols.append(f"{col}_{seen[col]}")
                            else:
                                seen[col] = 0
                                new_cols.append(col)
                                
                        src_df = raw_df.iloc[header_row_idx+1:].copy()
                        src_df.columns = new_cols
                        course_code_col = new_cols[course_code_col_idx]
                    else:
                        src_df = pd.read_excel(src_mapping_file, engine=engine)
                        course_code_col = None
                        for col in src_df.columns:
                            col_norm = re.sub(r'[^a-zA-Z0-9]', '', str(col)).upper()
                            if 'COURSECODE' in col_norm or 'SUBJCODE' in col_norm:
                                course_code_col = col
                                break
                        if course_code_col is None:
                            course_code_col = src_df.columns[0]
                    
                    # Normalize Course Code column values to search for target_code
                    match_idx = None
                    for idx, val in enumerate(src_df[course_code_col].values):
                        if pd.notna(val):
                            norm_val = re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()
                            # Try exact match first
                            if norm_val == target_code:
                                match_idx = idx
                                break
                            # Fallback: Compare core code (scheme-agnostic, e.g. "AI51" vs "AI51")
                            m_candidate = re.match(r'^\d+(.*)', norm_val)
                            candidate_core = m_candidate.group(1) if m_candidate else norm_val
                            if target_core and len(target_core) >= 3:
                                if target_core in candidate_core or candidate_core in target_core:
                                    match_idx = idx
                                    break
                                
                    if match_idx is not None:
                        # Collect sequential rows belonging to this course
                        co_rows = []
                        for idx in range(match_idx, len(src_df)):
                            val_code = src_df.iloc[idx][course_code_col]
                            if idx > match_idx and pd.notna(val_code) and str(val_code).strip() != '':
                                break
                            co_rows.append(src_df.iloc[idx])
                            
                        # Extract header columns mapping (e.g. 'PO1' -> Col 4)
                        header_cols = {}
                        for col in range(4, 19): # D to R
                            header_val = sheet_cppm.cell(row=3, column=col).value
                            if header_val:
                                header_cols[str(header_val).strip()] = col
                                
                        # Write the mapped values into the template
                        for co_idx, co_row in enumerate(co_rows):
                            target_row = 4 + co_idx # CO1 -> Row 4, CO2 -> Row 5...
                            if target_row > 13: # Limit to CO10
                                break
                                
                            co_row_keys_normalized = {re.sub(r'[^a-zA-Z0-9]', '', str(k)).upper(): k for k in co_row.keys()}
                            for col_name, col_idx in header_cols.items():
                                norm_col_name = re.sub(r'[^a-zA-Z0-9]', '', str(col_name)).upper()
                                if norm_col_name in co_row_keys_normalized:
                                    src_key = co_row_keys_normalized[norm_col_name]
                                    val = co_row[src_key]
                                    if pd.notna(val) and str(val).strip() != '':
                                        try:
                                            sheet_cppm.cell(row=target_row, column=col_idx).value = float(val)
                                        except ValueError:
                                            sheet_cppm.cell(row=target_row, column=col_idx).value = str(val)
                                    else:
                                        # Leave unmapped cells in active rows as empty/None or 0
                                        sheet_cppm.cell(row=target_row, column=col_idx).value = None
                except Exception as e:
                    # Log mapping warnings gracefully to Streamlit
                    st.warning(f"Note: Could not automatically map CO-PO weights from '{src_mapping_file}': {e}")
                
                
    out_stream = io.BytesIO()
    wb.save(out_stream)
    return out_stream.getvalue()

def process_stand_alone_lab(template_file, rubrics_file, api_key, co_vals, po_vals, course_code=None):
    """Fills the Stand Alone Lab sheet cleanly and handles dynamic rosters."""
    import google.generativeai as genai
    import fitz  # PyMuPDF
    from PIL import Image
    import io
    import json
    
    wb = load_workbook(io.BytesIO(template_file.read()))
    sheet = wb['Lab'] if 'Lab' in wb.sheetnames else wb.active
    
    # 1. Configure Course Type Cells in Theory sheet
    if 'Theory' in wb.sheetnames:
        sheet_theory = wb['Theory']
        sheet_theory.cell(row=7, column=4).value = "NO"  # STAND ALONE THEORY
        sheet_theory.cell(row=8, column=4).value = "YES" # STAND ALONE LAB
        sheet_theory.cell(row=9, column=4).value = "NO"  # IPCC
        if course_code:
            sheet_theory.cell(row=5, column=1).value = f"COURSE  CODE : {course_code}"
        
    # 2. Convert PDF pages to JPEG images in memory to optimize payload size
    images = []
    try:
        doc = fitz.open(stream=rubrics_file.read(), filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(dpi=150)
            img_bytes = pix.tobytes("jpg")
            images.append(img_bytes)
        doc.close()
    except Exception as e:
        raise ValueError(f"Failed to convert PDF to images: {e}")

    # 3. Configure Gemini API and extract marks
    if api_key:
        genai.configure(api_key=api_key)
    else:
        raise ValueError("Gemini API key is required. Add it to .env or type it in the sidebar.")
        
    prompt = """
    Analyze these scanned lab rubrics sheets containing handwritten student marks.
    The PDF contains four distinct tables, each spanning exactly two pages (total 8 pages).
    - Table 1: Pages 1 and 2
    - Table 2: Pages 3 and 4
    - Table 3: Pages 5 and 6
    - Table 4: Pages 7 and 8
    
    Each table may be split horizontally (meaning the same students appear on both pages of the table, with columns/experiments split across pages) or vertically (different students on each page).
    Your task is to extract and compile a consolidated profile for each student. Match students by their USN and Name across pages of the same table if they are split horizontally to get their full set of 12 marks.
    
    For each student, extract:
    1. USN (University Seat Number)
    2. Student Name
    3. The marks scored in each lab experiment (Lab 1 to Lab 12).
    
    CRITICAL INSTRUCTION FOR MARKS:
    - Under each lab experiment column (Lab 1 to Lab 12), there are multiple sub-columns (e.g. Performance, Viva, Total).
    - The ONLY score you should extract is the one under the sub-column labeled "T(30)" (Total out of 30) or simply "T".
    - Do NOT extract viva, write-up, or performance scores. Only extract the total mark under the "T(30)" or "T" column.
    - If a student was absent or has no marks for a specific lab, use null for that lab mark.
    
    Find the Course Code (e.g. 22AIL55, 21AIL35, etc.) printed at the top of the sheets.
    
    Return a structured JSON output with the following format:
    {
      "course_code": "21AIL35",
      "students": [
        {
          "usn": "1DS22AI001",
          "name": "ABHAY VIJAY GOUDAR",
          "marks": [28, 25, 30, 27, 29, 30, 28, 26, 29, 30, 27, 28]
        }
      ]
    }
    Make sure to match each student's marks correctly to their USN.
    """

    # Strict OpenAPI schema for Gemini VLM response format
    rubrics_schema = {
        "type": "OBJECT",
        "properties": {
            "course_code": {
                "type": "STRING",
                "description": "The course code printed at the top of the sheets, e.g. 21AIL35"
            },
            "students": {
                "type": "ARRAY",
                "items": {
                    "type": "OBJECT",
                    "properties": {
                        "usn": {
                            "type": "STRING",
                            "description": "University Seat Number (USN)"
                        },
                        "name": {
                            "type": "STRING",
                            "description": "Student Name"
                        },
                        "marks": {
                            "type": "ARRAY",
                            "items": {
                                "type": "NUMBER",
                                "nullable": True
                            },
                            "description": "Array of exactly 12 marks corresponding to Lab 1 through Lab 12. Use null if absent."
                        }
                    },
                    "required": ["usn", "name", "marks"]
                }
            }
        },
        "required": ["course_code", "students"]
    }

    # Try models in order of preference
    models_to_try = [
        'gemini-2.5-flash',
        'gemini-1.5-flash',
        'gemini-1.5-flash-8b',
        'gemini-2.0-flash',
        'gemini-1.5-pro',
        'gemini-2.0-flash-lite'
    ]
    
    try:
        available_models = [m.name.split('/')[-1] for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
        models_to_try = [m for m in models_to_try if m in available_models] or models_to_try
    except Exception as list_err:
        pass

    response = None
    last_error = None

    try:
        pil_images = [Image.open(io.BytesIO(img_bytes)) for img_bytes in images]
        
        for model_name in models_to_try:
            try:
                model = genai.GenerativeModel(model_name)
                response = model.generate_content(
                    [prompt, *pil_images],
                    generation_config={
                        "response_mime_type": "application/json",
                        "response_schema": rubrics_schema
                    }
                )
                # Success! Break the loop
                break
            except Exception as model_err:
                last_error = model_err
                err_msg = str(model_err).lower()
                if "429" in err_msg or "quota" in err_msg or "rate limit" in err_msg or "exhausted" in err_msg:
                    st.warning(f"⚠️ {model_name} quota exceeded or rate-limited. Retrying automatically with alternative model...")
                else:
                    st.warning(f"⚠️ {model_name} call failed: {model_err}. Retrying with alternative model...")
        
        if response is None:
            raise ValueError(f"All attempted models failed. Last error: {last_error}")
            
        data = json.loads(response.text)
        extracted_students = data.get("students", [])
        extracted_course_code = data.get("course_code", "")
        if extracted_course_code:
            course_code = str(extracted_course_code).strip()
            if 'Theory' in wb.sheetnames:
                sheet_theory = wb['Theory']
                sheet_theory.cell(row=5, column=1).value = f"COURSE  CODE : {course_code}"
    except Exception as e:
        raise ValueError(f"Gemini VLM API Call failed: {e}")

    # 4. Build Student Roster from VLM Extraction
    master_students = {}
    extracted_marks_map = {}
    for est in extracted_students:
        usn = str(est.get("usn", "")).strip()
        name = str(est.get("name", "")).strip()
        marks = est.get("marks", [])
        if usn and usn.lower() not in ('nan', 'none', '') and usn.isalnum():
            master_students[usn] = name
            extracted_marks_map[usn] = marks
            
    sorted_usns = sorted(master_students.keys())
    num_students = len(sorted_usns)
    
    if num_students == 0:
        raise ValueError("No students extracted from the scanned rubrics PDF.")

    # 5. Populate and Synchronize Roster to Lab sheet (Rows 9+) and Theory sheet (Rows 17 to 85)
    for i in range(num_students):
        usn = sorted_usns[i]
        name = master_students[usn]
        
        # Write to Lab sheet
        r_lab = 9 + i
        sheet.cell(row=r_lab, column=1).value = float(i + 1)
        sheet.cell(row=r_lab, column=2).value = usn
        sheet.cell(row=r_lab, column=3).value = name
        
        # Write to Theory sheet
        if 'Theory' in wb.sheetnames:
            sheet_theory = wb['Theory']
            r_theo = 17 + i
            sheet_theory.cell(row=r_theo, column=1).value = float(i + 1)
            sheet_theory.cell(row=r_theo, column=2).value = usn
            sheet_theory.cell(row=r_theo, column=3).value = name

    # Clear unused roster rows in Theory sheet (Rows 17 + N to 85)
    if 'Theory' in wb.sheetnames:
        sheet_theory = wb['Theory']
        for i in range(num_students, 85 - 17 + 1):
            r_theo = 17 + i
            for c in range(1, 123):
                sheet_theory.cell(row=r_theo, column=c).value = None

    # Clear student marks columns in Theory (columns D to CW, rows 17 to 17 + N - 1)
    if 'Theory' in wb.sheetnames:
        sheet_theory = wb['Theory']
        for r in range(17, 17 + num_students):
            for c in range(4, 102):
                sheet_theory.cell(row=r, column=c).value = None

    # 6. Fill COs and POs mapped in Lab sheet
    # Row 4 is POs Mapped, Row 5 is COs mapped. Columns D to O (4 to 15)
    for idx in range(12):
        col = 4 + idx
        if idx < len(po_vals) and po_vals[idx]:
            sheet.cell(row=4, column=col).value = po_vals[idx]
        if idx < len(co_vals) and co_vals[idx] is not None:
            sheet.cell(row=5, column=col).value = co_vals[idx]

    # Clear marks columns in Lab sheet (columns D to O, rows 9 to 9 + N - 1) before filling
    for r in range(9, 9 + num_students):
        for c in range(4, 16):
            sheet.cell(row=r, column=c).value = None

    # 7. Write Extracted Marks to Lab Sheet (columns D to O)
    for i in range(num_students):
        usn = sorted_usns[i]
        marks = extracted_marks_map[usn]
        r_lab = 9 + i
        for idx in range(12):
            col = 4 + idx
            if idx < len(marks):
                val = marks[idx]
                if val is not None and val != "":
                    try:
                        sheet.cell(row=r_lab, column=col).value = float(val)
                    except ValueError:
                        pass

    # 8. Prevent division by zero in other sheets by initializing unmapped values
    if 'CES' in wb.sheetnames:
        sheet_ces = wb['CES']
        for col in range(4, 14): # D to M
            if sheet_ces.cell(row=11, column=col).value is None:
                sheet_ces.cell(row=11, column=col).value = 0

    if 'CO_PO_PSO_MAPPING' in wb.sheetnames:
        sheet_cppm = wb['CO_PO_PSO_MAPPING']
        
        # Default initialization: fill all cells in CO1-CO10 (D4:R13) with 0 to prevent division by zero
        for r in range(4, 14):
            for c in range(4, 19): # Columns D to R
                sheet_cppm.cell(row=r, column=c).value = 0
        # Attempt to automatically load mapping based on course code
        if course_code:
            target_code = re.sub(r'[^a-zA-Z0-9]', '', str(course_code)).upper()
            
            # Extract core code by removing leading scheme indicators (e.g., "22", "21")
            m_target = re.match(r'^\d+(.*)', target_code)
            target_core = m_target.group(1) if m_target else target_code
            
            import os
            script_dir = os.path.dirname(os.path.abspath(__file__))
            src_mapping_file = None
            
            # Check current working directory first, then script directory
            for search_dir in ['.', script_dir]:
                try:
                    for f_name in os.listdir(search_dir):
                        f_upper = f_name.upper()
                        if 'CO' in f_upper and 'PO' in f_upper and 'MAPPING' in f_upper:
                            if f_upper.endswith('.XLSX') or f_upper.endswith('.XLS'):
                                src_mapping_file = os.path.join(search_dir, f_name)
                                break
                except:
                    pass
                if src_mapping_file:
                    break
                    
            if not src_mapping_file:
                for search_dir in ['.', script_dir]:
                    for fallback in ['2021-CO-PO -PSO-MAPPING.xlsx', '2021 CO PO PSO MAPPING.xls', '2021-CO-PO-PSO-MAPPING.xlsx', '2021-CO-PO-PSO-MAPPING.xls']:
                        fallback_path = os.path.join(search_dir, fallback)
                        if os.path.exists(fallback_path):
                            src_mapping_file = fallback_path
                            break
                    if src_mapping_file:
                        break
                
            if src_mapping_file:
                try:
                    engine = 'openpyxl' if src_mapping_file.lower().endswith('.xlsx') else 'xlrd'
                    raw_df = pd.read_excel(src_mapping_file, engine=engine, header=None)
                    
                    # Scan first 15 rows to find the header row containing 'Course Code'
                    header_row_idx = 0
                    course_code_col_idx = 0
                    found_header = False
                    
                    for r_idx in range(min(15, len(raw_df))):
                        row_vals = [str(x).strip().upper() for x in raw_df.iloc[r_idx].values]
                        for c_idx, val in enumerate(row_vals):
                            if ('COURSE' in val and 'CODE' in val) or ('SUBJ' in val and 'CODE' in val):
                                header_row_idx = r_idx
                                course_code_col_idx = c_idx
                                found_header = True
                                break
                        if found_header:
                            break
                            
                    if found_header:
                        cols = [str(x).strip() for x in raw_df.iloc[header_row_idx].values]
                        seen = {}
                        new_cols = []
                        for col in cols:
                            if col in seen:
                                seen[col] += 1
                                new_cols.append(f"{col}_{seen[col]}")
                            else:
                                seen[col] = 0
                                new_cols.append(col)
                                
                        src_df = raw_df.iloc[header_row_idx+1:].copy()
                        src_df.columns = new_cols
                        course_code_col = new_cols[course_code_col_idx]
                    else:
                        src_df = pd.read_excel(src_mapping_file, engine=engine)
                        course_code_col = None
                        for col in src_df.columns:
                            col_norm = re.sub(r'[^a-zA-Z0-9]', '', str(col)).upper()
                            if 'COURSECODE' in col_norm or 'SUBJCODE' in col_norm:
                                course_code_col = col
                                break
                        if course_code_col is None:
                            course_code_col = src_df.columns[0]
                    
                    # Normalize Course Code column values to search for target_code
                    match_idx = None
                    for idx, val in enumerate(src_df[course_code_col].values):
                        if pd.notna(val):
                            norm_val = re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()
                            # Try exact match first
                            if norm_val == target_code:
                                match_idx = idx
                                break
                            # Fallback: Compare core code (scheme-agnostic, e.g. "AI51" vs "AI51")
                            m_candidate = re.match(r'^\d+(.*)', norm_val)
                            candidate_core = m_candidate.group(1) if m_candidate else norm_val
                            if target_core and len(target_core) >= 3:
                                if target_core in candidate_core or candidate_core in target_core:
                                    match_idx = idx
                                    break
                                
                    if match_idx is not None:
                        # Collect sequential rows belonging to this course
                        co_rows = []
                        for idx in range(match_idx, len(src_df)):
                            val_code = src_df.iloc[idx][course_code_col]
                            if idx > match_idx and pd.notna(val_code) and str(val_code).strip() != '':
                                break
                            co_rows.append(src_df.iloc[idx])
                            
                        # Extract header columns mapping (e.g. 'PO1' -> Col 4)
                        header_cols = {}
                        for col in range(4, 19): # D to R
                            header_val = sheet_cppm.cell(row=3, column=col).value
                            if header_val:
                                header_cols[str(header_val).strip()] = col
                                
                        # Write the mapped values into the template
                        for co_idx, co_row in enumerate(co_rows):
                            target_row = 4 + co_idx # CO1 -> Row 4, CO2 -> Row 5...
                            if target_row > 13: # Limit to CO10
                                break
                                
                            co_row_keys_normalized = {re.sub(r'[^a-zA-Z0-9]', '', str(k)).upper(): k for k in co_row.keys()}
                            for col_name, col_idx in header_cols.items():
                                norm_col_name = re.sub(r'[^a-zA-Z0-9]', '', str(col_name)).upper()
                                if norm_col_name in co_row_keys_normalized:
                                    src_key = co_row_keys_normalized[norm_col_name]
                                    val = co_row[src_key]
                                    if pd.notna(val) and str(val).strip() != '':
                                        try:
                                            sheet_cppm.cell(row=target_row, column=col_idx).value = float(val)
                                        except ValueError:
                                            sheet_cppm.cell(row=target_row, column=col_idx).value = str(val)
                                    else:
                                        # Leave unmapped cells in active rows as empty/None or 0
                                        sheet_cppm.cell(row=target_row, column=col_idx).value = None
                except Exception as e:
                    # Log mapping warnings gracefully to Streamlit
                    st.warning(f"Note: Could not automatically map CO-PO weights from '{src_mapping_file}': {e}")

    # Clean unused rows in Lab sheet (from 9 + num_students to 81)
    for r in range(9 + num_students, 82):
        for c in range(1, sheet.max_column + 1):
            sheet.cell(row=r, column=c).value = None

    # Clean unused rows in CIA MARKS sheet (from 14 + num_students to 83)
    if 'CIA MARKS' in wb.sheetnames:
        sheet_cia = wb['CIA MARKS']
        for r in range(14 + num_students, 84):
            for c in range(1, sheet_cia.max_column + 1):
                sheet_cia.cell(row=r, column=c).value = None

    # Clear Quiz1 / Quiz2 / CES sheets for lab course to avoid division by zero
    for q_sheet_name in ['Quiz1', 'Quiz2', 'CES']:
        if q_sheet_name in wb.sheetnames:
            q_sheet = wb[q_sheet_name]
            for r in range(9, 74):
                for c in range(4, 80):
                    q_sheet.cell(row=r, column=c).value = None

    out_stream = io.BytesIO()
    wb.save(out_stream)
    return out_stream.getvalue()

def process_ipcc_course(template_file, qp_files, marks_files, quiz_file, aat_file, rubrics_file, api_key, co_vals, po_vals):
    """
    Consolidates both Theory assessments and Lab Rubrics VLM extraction
    into the IPCC template workbook (containing both Theory and Lab sheets).
    """
    import google.generativeai as genai
    import fitz  # PyMuPDF
    from PIL import Image
    import io
    import json

    wb = load_workbook(io.BytesIO(template_file.read()))
    sheet_theory = wb['Theory'] if 'Theory' in wb.sheetnames else wb.active
    sheet_lab = wb['Lab'] if 'Lab' in wb.sheetnames else wb.active
    
    # 1. Configure Course Type Cells in Theory sheet
    sheet_theory.cell(row=7, column=4).value = "NO"  # STAND ALONE THEORY
    sheet_theory.cell(row=8, column=4).value = "NO"  # STAND ALONE LAB
    sheet_theory.cell(row=9, column=4).value = "YES" # IPCC
    
    # 2. VLM Extraction if rubrics_file is provided (6 labs)
    extracted_students = []
    extracted_marks_map = {}
    course_code = None
    
    if rubrics_file and api_key:
        # Convert PDF pages to JPEG images in memory to optimize payload size
        images = []
        try:
            doc = fitz.open(stream=rubrics_file.read(), filetype="pdf")
            for page in doc:
                pix = page.get_pixmap(dpi=150)
                img_bytes = pix.tobytes("jpg")
                images.append(img_bytes)
            doc.close()
        except Exception as e:
            raise ValueError(f"Failed to convert PDF to images: {e}")

        # Configure Gemini API
        genai.configure(api_key=api_key)
        
        prompt = """
        Analyze these scanned lab rubrics sheets containing handwritten student marks.
        The PDF contains four distinct tables, each spanning exactly two pages (total 8 pages).
        - Table 1: Pages 1 and 2
        - Table 2: Pages 3 and 4
        - Table 3: Pages 5 and 6
        - Table 4: Pages 7 and 8
        
        Each table may be split horizontally (meaning the same students appear on both pages of the table, with columns/experiments split across pages) or vertically (different students on each page).
        Your task is to extract and compile a consolidated profile for each student. Match students by their USN and Name across pages of the same table if they are split horizontally to get their full set of 6 marks.
        
        For each student, extract:
        1. USN (University Seat Number)
        2. Student Name
        3. The marks scored in each lab experiment (Lab 1 to Lab 6).
        
        CRITICAL INSTRUCTION FOR MARKS:
        - Under each lab experiment column (Lab 1 to Lab 6), there are multiple sub-columns (e.g. Performance, Viva, Total).
        - The ONLY score you should extract is the one under the sub-column labeled "T(30)" (Total out of 30) or simply "T".
        - Do NOT extract viva, write-up, or performance scores. Only extract the total mark under the "T(30)" or "T" column.
        - If a student was absent or has no marks for a specific lab, use null for that lab mark.
        
        Find the Course Code (e.g. 22AIL55, 21AIL35, etc.) printed at the top of the sheets.
        
        Return a structured JSON output with the following format:
        {
          "course_code": "21AIL35",
          "students": [
            {
              "usn": "1DS22AI001",
              "name": "ABHAY VIJAY GOUDAR",
              "marks": [28, 25, 30, 27, 29, 30]
            }
          ]
        }
        Make sure to match each student's marks correctly to their USN.
        """

        # Strict OpenAPI schema for Gemini VLM response format (6 labs)
        rubrics_schema = {
            "type": "OBJECT",
            "properties": {
                "course_code": {
                    "type": "STRING",
                    "description": "The course code printed at the top of the sheets, e.g. 21AIL35"
                },
                "students": {
                    "type": "ARRAY",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "usn": {
                                "type": "STRING",
                                "description": "University Seat Number (USN)"
                            },
                            "name": {
                                "type": "STRING",
                                "description": "Student Name"
                            },
                            "marks": {
                                "type": "ARRAY",
                                "items": {
                                    "type": "NUMBER",
                                    "nullable": True
                                },
                                "description": "Array of exactly 6 marks corresponding to Lab 1 through Lab 6. Use null if absent."
                            }
                        },
                        "required": ["usn", "name", "marks"]
                    }
                }
            },
            "required": ["course_code", "students"]
        }

        # Try models in order of preference
        models_to_try = [
            'gemini-2.5-flash',
            'gemini-1.5-flash',
            'gemini-1.5-flash-8b',
            'gemini-2.0-flash',
            'gemini-1.5-pro',
            'gemini-2.0-flash-lite'
        ]
        
        try:
            available_models = [m.name.split('/')[-1] for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            models_to_try = [m for m in models_to_try if m in available_models] or models_to_try
        except Exception as list_err:
            pass

        response = None
        last_error = None

        try:
            pil_images = [Image.open(io.BytesIO(img_bytes)) for img_bytes in images]
            for model_name in models_to_try:
                try:
                    model = genai.GenerativeModel(model_name)
                    response = model.generate_content(
                        [prompt, *pil_images],
                        generation_config={
                            "response_mime_type": "application/json",
                            "response_schema": rubrics_schema
                        }
                    )
                    break
                except Exception as model_err:
                    last_error = model_err
                    err_msg = str(model_err).lower()
                    if "429" in err_msg or "quota" in err_msg or "rate limit" in err_msg or "exhausted" in err_msg:
                        st.warning(f"⚠️ {model_name} quota exceeded or rate-limited. Retrying automatically with alternative model...")
                    else:
                        st.warning(f"⚠️ {model_name} call failed: {model_err}. Retrying with alternative model...")
            
            if response is None:
                raise ValueError(f"All attempted models failed. Last error: {last_error}")
                
            data = json.loads(response.text)
            extracted_students = data.get("students", [])
            extracted_course_code = data.get("course_code", "")
            if extracted_course_code:
                course_code = str(extracted_course_code).strip()
        except Exception as e:
            raise ValueError(f"Gemini VLM API Call failed: {e}")

    # 3. Build Roster from both Theory & Lab VLM Extraction
    master_students = {}
    
    # A. Load students from Theory marks files
    all_marks_files = []
    if marks_files:
        all_marks_files.extend([f for f in marks_files.values() if f])
    if quiz_file:
        all_marks_files.append(quiz_file)
    if aat_file:
        all_marks_files.append(aat_file)
        
    for mf in all_marks_files:
        mf.seek(0)
        df = pd.read_excel(io.BytesIO(mf.read()), engine='xlrd', header=None)
        mf.seek(0)
        header_row_idx = None
        usn_col = None
        for idx, row in df.iterrows():
            row_vals = [str(x).strip() for x in row.values]
            if 'USN' in row_vals:
                header_row_idx = idx
                usn_col = row_vals.index('USN')
                break
        if header_row_idx is not None and usn_col is not None:
            row_vals = [str(x).strip() for x in df.iloc[header_row_idx].values]
            name_col = None
            for c_idx, val in enumerate(row_vals):
                if 'student' in val.lower() or 'name' in val.lower():
                    if 'staff' not in val.lower() and 'division' not in val.lower():
                        name_col = c_idx
                        break
            if name_col is None:
                name_col = usn_col + 1
            for idx in range(header_row_idx + 1, len(df)):
                row = df.iloc[idx]
                u = str(row.iloc[usn_col]).strip()
                n = str(row.iloc[name_col]).strip() if name_col < len(row) else ""
                if u and u.lower() not in ('nan', 'none', '') and u.isalnum():
                    master_students[u] = n

    # B. Load students from Lab VLM extraction
    for est in extracted_students:
        usn = str(est.get("usn", "")).strip()
        name = str(est.get("name", "")).strip()
        marks = est.get("marks", [])
        if usn and usn.lower() not in ('nan', 'none', '') and usn.isalnum():
            master_students[usn] = name
            extracted_marks_map[usn] = marks

    sorted_usns = sorted(master_students.keys())
    num_students = len(sorted_usns)
    
    if num_students == 0:
        raise ValueError("No students found in Theory files or Lab VLM extraction.")

    # 4. Populate and Synchronize Roster to Lab sheet (Rows 9+) and Theory sheet (Rows 17 to 85)
    for i in range(85 - 17 + 1):
        r_theo = 17 + i
        if i < num_students:
            usn = sorted_usns[i]
            name = master_students[usn]
            sheet_theory.cell(row=r_theo, column=1).value = float(i + 1)
            sheet_theory.cell(row=r_theo, column=2).value = usn
            sheet_theory.cell(row=r_theo, column=3).value = name
        else:
            for c in range(1, 123):
                sheet_theory.cell(row=r_theo, column=c).value = None

    # Clear Theory marks columns (columns D to CW, rows 17 to 17 + N - 1)
    for r in range(17, 17 + num_students):
        for c in range(4, 102):
            sheet_theory.cell(row=r, column=c).value = None

    # Write Roster to Lab sheet and Sum Formulas in Column 10 (J)
    for i in range(num_students):
        usn = sorted_usns[i]
        name = master_students[usn]
        r_lab = 9 + i
        sheet_lab.cell(row=r_lab, column=1).value = float(i + 1)
        sheet_lab.cell(row=r_lab, column=2).value = usn
        sheet_lab.cell(row=r_lab, column=3).value = name
        sheet_lab.cell(row=r_lab, column=10).value = f"=SUM(D{r_lab}:I{r_lab})"

    # Clear unused roster rows in Lab sheet
    for r in range(9 + num_students, 82):
        for c in range(1, sheet_lab.max_column + 1):
            sheet_lab.cell(row=r, column=c).value = None

    # Clear unused student rows in CIA MARKS sheet
    if 'CIA MARKS' in wb.sheetnames:
        sheet_cia = wb['CIA MARKS']
        for r in range(14 + num_students, 84):
            for c in range(1, sheet_cia.max_column + 1):
                sheet_cia.cell(row=r, column=c).value = None

    # 5. Extract and write metadata from IA1 Marks if available
    metadata = {}
    for test_idx in [1, 2, 3]:
        if marks_files.get(test_idx):
            marks_files[test_idx].seek(0)
            metadata = extract_metadata(marks_files[test_idx])
            marks_files[test_idx].seek(0)
            if metadata['course_code']:
                break
                
    if course_code:
        metadata['course_code'] = course_code
        
    if metadata:
        if metadata['academic_year']:
            sheet_theory.cell(row=4, column=1).value = f"Academic Year : {metadata['academic_year']}"
        if metadata['batch']:
            sheet_theory.cell(row=4, column=4).value = f"Batch :{metadata['batch']}"
        if metadata['semester']:
            sheet_theory.cell(row=4, column=17).value = f"Semester  : {metadata['semester']}"
        if metadata['course_code']:
            sheet_theory.cell(row=5, column=1).value = f"COURSE  CODE : {metadata['course_code']}"
        if metadata['course_title']:
            sheet_theory.cell(row=6, column=1).value = f"COURSE TITLE : {metadata['course_title']}"
        if metadata['faculty']:
            sheet_theory.cell(row=10, column=1).value = f"Faculty : {metadata['faculty']}"

    # Map Test Columns in Theory sheet
    tests_col_map = {1: {}, 2: {}, 3: {}} 
    first_q1a_col = -1
    for col in range(1, 200):
        val = sheet_theory.cell(row=16, column=col).value
        val_str = str(val).strip() if val else ""
        if val_str == 'Q1A':
            first_q1a_col = col
            break

    if first_q1a_col != -1:
        for test_idx in [1, 2, 3]:
            start_col = first_q1a_col + (test_idx - 1) * 32
            for q_num in range(1, 9):
                tests_col_map[test_idx][q_num] = start_col + (q_num - 1) * 4

    # Write Test COs and Marks in Theory
    for test_idx in [1, 2, 3]:
        qp = qp_files.get(test_idx)
        marks = marks_files.get(test_idx)
        col_mappings = tests_col_map[test_idx]
        
        # Clear Row 12, 14, and 15
        for col_idx in col_mappings.values():
            for offset in range(4):
                sheet_theory.cell(row=12, column=col_idx + offset).value = None
                sheet_theory.cell(row=14, column=col_idx + offset).value = None
                sheet_theory.cell(row=15, column=col_idx + offset).value = None
        
        # CO Mappings
        if qp:
            qp.seek(0)
            co_map = extract_cos_from_docx(qp)
            qp.seek(0)
            part_offsets = {'a': 0, 'b': 1, 'c': 2, 'd': 3}
            for (q_num, part), co_val in co_map.items():
                if q_num in col_mappings:
                    col_start = col_mappings[q_num]
                    target_col = col_start + part_offsets.get(part, 0)
                    sheet_theory.cell(row=12, column=target_col).value = f"0,{co_val}"
                    
        # Marks Mappings
        if marks:
            marks.seek(0)
            st_marks = extract_marks_from_xls(marks)
            marks.seek(0)
            part_offsets = {'a': 0, 'b': 1, 'c': 2, 'd': 3}
            for i in range(num_students):
                r = 17 + i
                usn_val = sheet_theory.cell(row=r, column=2).value
                if usn_val in st_marks:
                    for (q_num, part), mark in st_marks[usn_val].items():
                        if q_num in col_mappings:
                            col_start = col_mappings[q_num]
                            target_col = col_start + part_offsets.get(part, 0)
                            sheet_theory.cell(row=r, column=target_col).value = mark

        # Set maximum marks and thresholds automatically based on part B activity
        for q_num in range(1, 9):
            col_a = col_mappings[q_num]
            col_b = col_mappings[q_num] + 1
            part_b_active = False
            for i in range(num_students):
                r = 17 + i
                if sheet_theory.cell(row=r, column=col_b).value is not None:
                    part_b_active = True
                    break
            if sheet_theory.cell(row=12, column=col_b).value is not None:
                part_b_active = True
                
            if part_b_active:
                sheet_theory.cell(row=14, column=col_a).value = 5.0
                sheet_theory.cell(row=14, column=col_b).value = 5.0
                col_a_let = get_column_letter(col_a)
                col_b_let = get_column_letter(col_b)
                sheet_theory.cell(row=15, column=col_a).value = f"=PRODUCT({col_a_let}14,0.65)"
                sheet_theory.cell(row=15, column=col_b).value = f"=PRODUCT({col_b_let}14,0.65)"
            else:
                sheet_theory.cell(row=14, column=col_a).value = 10.0
                col_a_let = get_column_letter(col_a)
                sheet_theory.cell(row=15, column=col_a).value = f"=PRODUCT({col_a_let}14,0.65)"

    # Map Quiz & AAT Marks
    if quiz_file:
        quiz_file.seek(0)
        quiz_marks = extract_single_column_marks(quiz_file)
        quiz_file.seek(0)
        for i in range(num_students):
            r = 17 + i
            usn_val = sheet_theory.cell(row=r, column=2).value
            if usn_val in quiz_marks:
                sheet_theory.cell(row=r, column=101).value = quiz_marks[usn_val]
                
    if aat_file:
        aat_file.seek(0)
        aat_marks = extract_single_column_marks(aat_file)
        aat_file.seek(0)
        for i in range(num_students):
            r = 17 + i
            usn_val = sheet_theory.cell(row=r, column=2).value
            if usn_val in aat_marks:
                sheet_theory.cell(row=r, column=100).value = aat_marks[usn_val]

    # 6. Write Lab Assessment Marks & CO/PO Mappings (6 labs)
    # Fill COs and POs mapped in Lab sheet (Row 4 is POs, Row 5 is COs, Cols D to I)
    for idx in range(6):
        col = 4 + idx
        if idx < len(po_vals) and po_vals[idx]:
            sheet_lab.cell(row=4, column=col).value = po_vals[idx]
        if idx < len(co_vals) and co_vals[idx] is not None:
            sheet_lab.cell(row=5, column=col).value = co_vals[idx]

    # Clear marks columns in Lab (cols D to I, rows 9 to 9 + N - 1) before writing
    for r in range(9, 9 + num_students):
        for c in range(4, 10): # Cols D to I (4 to 9)
            sheet_lab.cell(row=r, column=c).value = None

    # Write VLM extracted marks
    for i in range(num_students):
        usn = sorted_usns[i]
        r_lab = 9 + i
        if usn in extracted_marks_map:
            marks = extracted_marks_map[usn]
            for idx in range(6):
                col = 4 + idx
                if idx < len(marks):
                    val = marks[idx]
                    if val is not None and val != "":
                        try:
                            sheet_lab.cell(row=r_lab, column=col).value = float(val)
                        except ValueError:
                            pass

    # 7. Protection / SEE Grade "NA"
    for i in range(num_students):
        r = 17 + i
        sheet_theory.cell(row=r, column=108).value = "NA"

    # CES Sheet adjustments
    if 'CES' in wb.sheetnames:
        sheet_ces = wb['CES']
        for col in range(4, 14):
            if sheet_ces.cell(row=11, column=col).value is None:
                sheet_ces.cell(row=11, column=col).value = 0

    # Map CO-PO-PSO Matrix
    if 'CO_PO_PSO_MAPPING' in wb.sheetnames:
        sheet_cppm = wb['CO_PO_PSO_MAPPING']
        for r in range(4, 14):
            for c in range(4, 19):
                sheet_cppm.cell(row=r, column=c).value = 0

        ccode = metadata.get('course_code')
        if ccode:
            target_code = re.sub(r'[^a-zA-Z0-9]', '', str(ccode)).upper()
            m_target = re.match(r'^\d+(.*)', target_code)
            target_core = m_target.group(1) if m_target else target_code
            
            import os
            script_dir = os.path.dirname(os.path.abspath(__file__))
            src_mapping_file = None
            for search_dir in ['.', script_dir]:
                try:
                    for f_name in os.listdir(search_dir):
                        f_upper = f_name.upper()
                        if 'CO' in f_upper and 'PO' in f_upper and 'MAPPING' in f_upper:
                            if f_upper.endswith('.XLSX') or f_upper.endswith('.XLS'):
                                src_mapping_file = os.path.join(search_dir, f_name)
                                break
                except:
                    pass
                if src_mapping_file:
                    break
            if not src_mapping_file:
                for search_dir in ['.', script_dir]:
                    for fallback in ['2021-CO-PO -PSO-MAPPING.xlsx', '2021 CO PO PSO MAPPING.xls', '2021-CO-PO-PSO-MAPPING.xlsx', '2021-CO-PO-PSO-MAPPING.xls']:
                        fallback_path = os.path.join(search_dir, fallback)
                        if os.path.exists(fallback_path):
                            src_mapping_file = fallback_path
                            break
                    if src_mapping_file:
                        break
            if src_mapping_file:
                try:
                    engine = 'openpyxl' if src_mapping_file.lower().endswith('.xlsx') else 'xlrd'
                    raw_df = pd.read_excel(src_mapping_file, engine=engine, header=None)
                    header_row_idx = 0
                    course_code_col_idx = 0
                    found_header = False
                    for r_idx in range(min(15, len(raw_df))):
                        row_vals = [str(x).strip().upper() for x in raw_df.iloc[r_idx].values]
                        for c_idx, val in enumerate(row_vals):
                            if ('COURSE' in val and 'CODE' in val) or ('SUBJ' in val and 'CODE' in val):
                                header_row_idx = r_idx
                                course_code_col_idx = c_idx
                                found_header = True
                                break
                        if found_header:
                            break
                    if found_header:
                        cols = [str(x).strip() for x in raw_df.iloc[header_row_idx].values]
                        seen = {}
                        new_cols = []
                        for col in cols:
                            if col in seen:
                                seen[col] += 1
                                new_cols.append(f"{col}_{seen[col]}")
                            else:
                                seen[col] = 0
                                new_cols.append(col)
                        src_df = raw_df.iloc[header_row_idx+1:].copy()
                        src_df.columns = new_cols
                        course_code_col = new_cols[course_code_col_idx]
                    else:
                        src_df = pd.read_excel(src_mapping_file, engine=engine)
                        course_code_col = None
                        for col in src_df.columns:
                            col_norm = re.sub(r'[^a-zA-Z0-9]', '', str(col)).upper()
                            if 'COURSECODE' in col_norm or 'SUBJCODE' in col_norm:
                                course_code_col = col
                                break
                        if course_code_col is None:
                            course_code_col = src_df.columns[0]
                    
                    match_idx = None
                    for idx, val in enumerate(src_df[course_code_col].values):
                        if pd.notna(val):
                            norm_val = re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()
                            if norm_val == target_code:
                                match_idx = idx
                                break
                            m_candidate = re.match(r'^\d+(.*)', norm_val)
                            candidate_core = m_candidate.group(1) if m_candidate else norm_val
                            if target_core and len(target_core) >= 3:
                                if target_core in candidate_core or candidate_core in target_core:
                                    match_idx = idx
                                    break
                    if match_idx is not None:
                        co_rows = []
                        for idx in range(match_idx, len(src_df)):
                            val_code = src_df.iloc[idx][course_code_col]
                            if idx > match_idx and pd.notna(val_code) and str(val_code).strip() != '':
                                break
                            co_rows.append(src_df.iloc[idx])
                        header_cols = {}
                        for col in range(4, 19):
                            header_val = sheet_cppm.cell(row=3, column=col).value
                            if header_val:
                                header_cols[str(header_val).strip()] = col
                        for co_idx, co_row in enumerate(co_rows):
                            target_row = 4 + co_idx
                            if target_row > 13:
                                break
                            co_row_keys_normalized = {re.sub(r'[^a-zA-Z0-9]', '', str(k)).upper(): k for k in co_row.keys()}
                            for col_name, col_idx in header_cols.items():
                                norm_col_name = re.sub(r'[^a-zA-Z0-9]', '', str(col_name)).upper()
                                if norm_col_name in co_row_keys_normalized:
                                    src_key = co_row_keys_normalized[norm_col_name]
                                    val = co_row[src_key]
                                    if pd.notna(val) and str(val).strip() != '':
                                        try:
                                            sheet_cppm.cell(row=target_row, column=col_idx).value = float(val)
                                        except ValueError:
                                            sheet_cppm.cell(row=target_row, column=col_idx).value = str(val)
                                    else:
                                        sheet_cppm.cell(row=target_row, column=col_idx).value = None
                except Exception as e:
                    st.warning(f"Note: Could not automatically map CO-PO weights from '{src_mapping_file}': {e}")

    # Clear Quiz1 / Quiz2 sheets
    for q_sheet_name in ['Quiz1', 'Quiz2']:
        if q_sheet_name in wb.sheetnames:
            q_sheet = wb[q_sheet_name]
            for r in range(9, 74):
                for c in range(4, 80):
                    q_sheet.cell(row=r, column=c).value = None

    out_stream = io.BytesIO()
    wb.save(out_stream)
    return out_stream.getvalue()

# ----------------- STREAMLIT INTERFACE -----------------

st.markdown('<div class="main-title">📊 CIA Marks & CO Mapper</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Seamlessly automate Course Outcome mapping and marks consolidation.</div>', unsafe_allow_html=True)

# Configuration Panel
st.sidebar.markdown('## 🛠️ Course Selection')
course_type = st.sidebar.radio(
    "Select Course Component",
    ["Theory Course", "Lab Course", "IPCC Course"],
    help="Choose the component type you are processing."
)

st.header("1. Choose Course Configuration")

col_type1, col_type2, col_type3 = st.columns(3)

with col_type1:
    theory_active = '<span class="tag-active">Active</span>' if course_type == "Theory Course" else ''
    theory_opacity = '1.0' if course_type == "Theory Course" else '0.6'
    st.markdown(f"""
    <div class="card" style="opacity: {theory_opacity};">
        <h3>Theory Course {theory_active}</h3>
        <p>Map CIA Question Papers, Student Marks, AAT, and Quiz into the Stand Alone Theory template.</p>
    </div>
    """, unsafe_allow_html=True)
    
with col_type2:
    lab_active = '<span class="tag-active">Active</span>' if course_type == "Lab Course" else ''
    lab_opacity = '1.0' if course_type == "Lab Course" else '0.6'
    st.markdown(f"""
    <div class="card" style="opacity: {lab_opacity};">
        <h3>Lab Course {lab_active}</h3>
        <p>Consolidate laboratory component evaluations and mappings cleanly using Vision VLM API.</p>
    </div>
    """, unsafe_allow_html=True)

with col_type3:
    ipcc_active = '<span class="tag-active">Active</span>' if course_type == "IPCC Course" else ''
    ipcc_opacity = '1.0' if course_type == "IPCC Course" else '0.6'
    st.markdown(f"""
    <div class="card" style="opacity: {ipcc_opacity};">
        <h3>IPCC Course {ipcc_active}</h3>
        <p>Integrated Professional Core Course mapping combining both Theory and Practical slots.</p>
    </div>
    """, unsafe_allow_html=True)

# Sidebar - Template Upload
with st.sidebar:
    st.markdown('<h2 style="font-weight: 800;">⚙️ Setup Template</h2>', unsafe_allow_html=True)
    if course_type == "Theory Course":
        template_file = st.file_uploader(
            "Upload Stand Alone Theory Template (.xlsx)",
            type=["xlsx"],
            help="Provide the empty STAND ALONE THEORY EMPTY TEMPLATE.xlsx workbook."
        )
    elif course_type == "Lab Course":
        template_file = st.file_uploader(
            "Upload Stand Alone Lab Template (.xlsx)",
            type=["xlsx"],
            help="Provide the empty STAND ALONE LAB-empty.xlsx workbook."
        )
    else:
        template_file = st.file_uploader(
            "Upload IPCC Template (.xlsx)",
            type=["xlsx"],
            help="Provide the empty IPCC-EMPTY TEMPLATE.xlsx workbook."
        )
    st.markdown("---")
    st.markdown("<small>Designed for department course mapping operations.</small>", unsafe_allow_html=True)

st.header("2. Upload Assessment Data")

if course_type in ("Theory Course", "IPCC Course"):
    # Render tabs dynamically
    if course_type == "Theory Course":
        tabs_list = ["📌 CIA-1", "📌 CIA-2", "📌 CIA-3", "📌 Quiz & AAT"]
    else:
        tabs_list = ["📌 CIA-1", "📌 CIA-2", "📌 CIA-3", "📌 Quiz & AAT", "📌 Lab Component"]
        
    tabs = st.tabs(tabs_list)

    qp_files = {}
    marks_files = {}
    quiz_file = None
    aat_file = None
    
    # Lab defaults (IPCC)
    co_input = "1, 1, 1, 2, 2, 2"
    po_input = "1,5; 1,5; 1,5; 2,5; 2,5; 2,5"
    rubrics_file = None
    user_api_key = ""

    with tabs[0]:
        col1, col2 = st.columns(2)
        with col1:
            qp_files[1] = st.file_uploader("CIA-1 Question Paper (.docx)", type=["docx"], key="qp1")
        with col2:
            marks_files[1] = st.file_uploader("CIA-1 IA Marks (.xls)", type=["xls"], key="m1")

    with tabs[1]:
        col1, col2 = st.columns(2)
        with col1:
            qp_files[2] = st.file_uploader("CIA-2 Question Paper (.docx)", type=["docx"], key="qp2")
        with col2:
            marks_files[2] = st.file_uploader("CIA-2 IA Marks (.xls)", type=["xls"], key="m2")

    with tabs[2]:
        col1, col2 = st.columns(2)
        with col1:
            qp_files[3] = st.file_uploader("CIA-3 Question Paper (.docx)", type=["docx"], key="qp3")
        with col2:
            marks_files[3] = st.file_uploader("CIA-3 IA Marks (.xls)", type=["xls"], key="m3")

    with tabs[3]:
        col1, col2 = st.columns(2)
        with col1:
            quiz_file = st.file_uploader("Quiz Marks (.xls)", type=["xls"], key="quiz")
        with col2:
            aat_file = st.file_uploader("AAT Marks (.xls)", type=["xls"], key="aat")
            
    if course_type == "IPCC Course":
        with tabs[4]:
            st.subheader("Configure Lab Mapping Parameters")
            col_co, col_po = st.columns(2)
            with col_co:
                co_input = st.text_input(
                    "Lab CO Mappings (Row 5)", 
                    value="1, 1, 1, 2, 2, 2",
                    help="Enter exactly 6 CO values separated by commas for columns D to I."
                )
            with col_po:
                po_input = st.text_input(
                    "Lab PO Mappings (Row 4)", 
                    value="1,5; 1,5; 1,5; 2,5; 2,5; 2,5",
                    help="Enter PO mappings for each lab. Separate labs with a semicolon (;) and PO numbers with commas (,)."
                )
                
            st.subheader("Upload Rubrics Scanned PDF")
            rubrics_file = st.file_uploader(
                "Upload Scanned Lab Rubrics PDF (.pdf)",
                type=["pdf"],
                help="Scanned PDF showing student marks for all 6 labs."
            )
            
            st.subheader("Gemini Vision API Configuration")
            env_key = os.environ.get("GEMINI_API_KEY") or ""
            user_api_key = st.text_input(
                "Gemini API Key Override (Optional)",
                value="",
                type="password",
                help="If not set in .env, paste your key here. Leave blank to use the key configured in the .env file."
            )

else:
    # Lab Course Upload Section
    st.subheader("Configure Lab Mapping Parameters")
    col_co, col_po = st.columns(2)
    with col_co:
        co_input = st.text_input(
            "Lab CO Mappings (Row 5)", 
            value="1, 1, 1, 2, 2, 2, 3, 3, 3, 4, 4, 4",
            help="Enter exactly 12 CO values separated by commas for columns D to O."
        )
    with col_po:
        po_input = st.text_input(
            "Lab PO Mappings (Row 4)", 
            value="1,5; 1,5; 1,5; 2,5; 2,5; 2,5; 3,5; 3,5; 3,5; 4,5; 4,5; 4,5",
            help="Enter PO mappings for each lab. Separate labs with a semicolon (;) and PO numbers with commas (,). We format it automatically."
        )
        
    st.subheader("Upload Rubrics Scanned PDF")
    rubrics_file = st.file_uploader(
        "Upload Scanned Lab Rubrics PDF (.pdf)",
        type=["pdf"],
        help="Scanned PDF showing student marks for all 12 labs."
    )
    
    st.subheader("Gemini Vision API Configuration")
    env_key = os.environ.get("GEMINI_API_KEY") or ""
    user_api_key = st.text_input(
        "Gemini API Key Override (Optional)",
        value="",
        type="password",
        help="If not set in .env, paste your key here. Leave blank to use the key configured in the .env file."
    )

st.markdown("<br>", unsafe_allow_html=True)

if st.button("Generate Consolidated Excel", type="primary"):
    if not template_file:
        st.error("⚠️ Please configure the Excel Template in the sidebar before processing.")
    elif course_type == "Theory Course" and not any(marks_files.values()) and not quiz_file and not aat_file:
        st.error("⚠️ Please upload at least one marks file (CIA, Quiz, or AAT) to map scores.")
    elif course_type == "Lab Course" and not rubrics_file:
        st.error("⚠️ Please upload the scanned rubrics PDF containing handwritten marks.")
    elif course_type == "IPCC Course" and not any(marks_files.values()) and not quiz_file and not aat_file and not rubrics_file:
        st.error("⚠️ Please upload at least one assessment data file (Theory or Lab Rubrics) to process.")
    else:
        with st.spinner("Processing templates and compiling rosters..."):
            try:
                if course_type == "Theory Course":
                    out_bytes = process_stand_alone_theory(
                        template_file,
                        qp_files,
                        marks_files,
                        quiz_file,
                        aat_file
                    )
                    filename_suggest = "Consolidated_Theory_Scheme.xlsx"
                elif course_type == "Lab Course":
                    # Parse CO values
                    co_vals = []
                    for val in co_input.split(","):
                        val_str = val.strip()
                        if val_str:
                            try:
                                co_vals.append(float(val_str))
                            except ValueError:
                                co_vals.append(val_str)
                    
                    if len(co_vals) != 12:
                        st.warning(f"Note: You entered {len(co_vals)} CO values. Padded with defaults to make 12.")
                        co_vals = (co_vals + [1.0]*12)[:12]

                    # Parse PO values
                    raw_po_groups = po_input.split(";")
                    po_vals = []
                    for grp in raw_po_groups:
                        grp = grp.strip()
                        if grp:
                            digits = [d.strip() for d in grp.split(",") if d.strip().isdigit()]
                            if digits:
                                po_vals.append("0," + ",".join(digits) + ",0")
                            else:
                                po_vals.append("")
                        else:
                            po_vals.append("")
                    
                    if len(po_vals) != 12:
                        st.warning(f"Note: You entered {len(po_vals)} PO groups. Padded with defaults to make 12.")
                        po_vals = (po_vals + [""]*12)[:12]

                    api_key = user_api_key if user_api_key else env_key
                    if not api_key:
                        raise ValueError("No Gemini API key found. Please define GEMINI_API_KEY in your .env file or input it above.")

                    out_bytes = process_stand_alone_lab(
                        template_file,
                        rubrics_file,
                        api_key,
                        co_vals,
                        po_vals
                    )
                    filename_suggest = "Consolidated_Lab_Scheme.xlsx"
                else:
                    # IPCC Course
                    # Parse CO values (6 values)
                    co_vals = []
                    for val in co_input.split(","):
                        val_str = val.strip()
                        if val_str:
                            try:
                                co_vals.append(float(val_str))
                            except ValueError:
                                co_vals.append(val_str)
                    if len(co_vals) != 6:
                        st.warning(f"Note: You entered {len(co_vals)} CO values. Padded with defaults to make 6.")
                        co_vals = (co_vals + [1.0]*6)[:6]

                    # Parse PO values (6 groups)
                    raw_po_groups = po_input.split(";")
                    po_vals = []
                    for grp in raw_po_groups:
                        grp = grp.strip()
                        if grp:
                            digits = [d.strip() for d in grp.split(",") if d.strip().isdigit()]
                            if digits:
                                po_vals.append("0," + ",".join(digits) + ",0")
                            else:
                                po_vals.append("")
                        else:
                            po_vals.append("")
                    if len(po_vals) != 6:
                        st.warning(f"Note: You entered {len(po_vals)} PO groups. Padded with defaults to make 6.")
                        po_vals = (po_vals + [""]*6)[:6]

                    api_key = user_api_key if user_api_key else env_key
                    if rubrics_file and not api_key:
                        raise ValueError("No Gemini API key found. Please define GEMINI_API_KEY in your .env file or input it above to process the Lab PDF.")

                    out_bytes = process_ipcc_course(
                        template_file,
                        qp_files,
                        marks_files,
                        quiz_file,
                        aat_file,
                        rubrics_file,
                        api_key,
                        co_vals,
                        po_vals
                    )
                    filename_suggest = "Consolidated_IPCC_Scheme.xlsx"

                st.success("🎉 Successfully compiled course data and generated spreadsheet!")
                st.download_button(
                    label="📥 Download Consolidated Scheme (.xlsx)",
                    data=out_bytes,
                    file_name=filename_suggest,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ An error occurred during mapping: {str(e)}")
                st.info("Check that all files match the expected formats and that your Gemini API key is active.")
